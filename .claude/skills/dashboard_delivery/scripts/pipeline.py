# -*- coding: utf-8 -*-
"""
Pipeline de /dashboard_delivery — ingesta del reporte mensual del PM (y de backfills
históricos puntuales) y regeneración del dashboard "Performance de desarrollo".

Uso (desde cualquier cwd):
    python pipeline.py inspect   # analiza el/los Excel de raw/ sin escribir nada
    python pipeline.py ingest    # ingesta completa: merge al log + regenera dashboard

Fuentes de verdad acumuladas (nunca se releen Excels históricos):
  - wiki/3_recursos/datos/log_performance_desarrollo.md  (tickets/SP publicados en producción)
  - wiki/3_recursos/datos/log_costos_desarrollo.md       (USD gastados en desarrollo, stock de
    horas mensual que factura Fintexa)
  - wiki/3_recursos/datos/log_sla_highest.md             (SLA de tickets Highest)

CONTRATO DE ESCRITURA (pipeline multi-PM, 2026-08-15) — este script YA NO escribe el canon
directo. Las rutas de arriba son un ESPEJO read-only en este install; el único que las
escribe es `/context_merge`, sobre el repo compartido CEREBRO_CORE. `ingest` siembra una
copia de trabajo en `wiki/1_proyectos/contexto_vivo/_staging_dashboard_delivery/` a partir
del espejo (o de un item `tipo: dato` pendiente sin mergear, si lo hay — verificar a mano)
y el resto del pipeline sigue leyendo/escribiendo esos 3 logs exactamente igual que antes,
solo que las constantes ahora apuntan ahí. Al terminar, la skill empaqueta esa carpeta como
UN item `tipo: dato` en `contexto_vivo/` con `destino_propuesto: 3_recursos/datos/` — el
merge lo aplica por copia byte a byte. El dashboard HTML (`outputs/`) no es canon, sigue
escribiéndose directo. Ver SKILL.md, Paso de cierre.

Formatos de origen reconocidos (auto-detectados por header, ver sniff_and_read):
  1. Ticket-level (export Jira del PM, mensual): una fila por ticket, con columna
     "Clave de incidencia" y "Mes". Año no viene en el archivo -> se asume ANIO_DEFAULT.
  2. Agregado por versión (backfills históricos encontrados, ej. pre-2026): una fila
     por versión publicada, con columnas SP-US/SP-BUGS/Q-US/Q-BUGS y AÑO PUBLICACIÓN.
     Sin Epic -> los registros quedan con epic=None (sentinel "sin dato de Epic",
     NO "(sin epic)": el dashboard oculta esos meses en las vistas por Epic en vez de
     agruparlos en un bucket gris, tal como pidió el usuario 2026-07-14).
  3. Stock de horas mensual (factura de Fintexa): una fila por recurso asignado, con
     columnas "Componente / Proyecto" y "Horas Mes", hoja titulada "Stock de Horas -
     <MES><AA>". Alimenta el log de COSTOS, no el de delivery (decisión del usuario
     2026-07-21). Ver detalle de reglas en la sección correspondiente más abajo.
  4. SLA de tickets Highest (export Jira "tiempo por estado", CSV, sesión 2026-07-27):
     una fila por ticket Prioridad=Highest, con columna "Clave" + una columna de tiempo
     acumulado por cada estado del workflow ("Asignado", "Backlog"*, "Bloqueado", "Con
     defecto", "Finalizada", "En curso", "EN QA", etc. — mismo formato "tiempo en
     estado" que usa /dashboard_qa, no fechas de transición). Alimenta el log de SLA
     (`log_sla_highest.md`), no el de delivery ni el de costos — mide cuánto tarda el
     equipo en resolver urgencias, no qué ni cuánto se publicó.
     **Fechas reales de versión (Jira, no adivinadas):** este formato NO trae fecha de
     release por versión, así que antes de ingerir hay que consultar Jira (MCP) por el
     campo `fixVersions`/`releaseDate` de cada ticket Finalizada del CSV, y volcarlo a
     un JSON compañero `<mismo nombre del csv>.versions.json` en `raw/` (mapa
     Clave -> [fechas ISO de versiones released]). El pipeline lo lee si existe; si no
     hay versión released con fecha para un ticket, degrada a la fecha proxy de entrada
     a "Finalizada" (`fuente_publicacion: "proxy_finalizada"`) y lo advierte.
"""
import csv
import json
import re
import shutil
import statistics
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

# --- Paths ------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
REPO = SCRIPT_DIR.parents[3]  # .claude/skills/dashboard_delivery/scripts -> repo root
RAW_DIR = REPO / "raw"

# Espejo read-only del canon (input) -- NUNCA se escribe acá.
MIRROR_LOG_MD = REPO / "wiki" / "3_recursos" / "datos" / "log_performance_desarrollo.md"
MIRROR_LOG_COST_MD = REPO / "wiki" / "3_recursos" / "datos" / "log_costos_desarrollo.md"
MIRROR_LOG_SLA_MD = REPO / "wiki" / "3_recursos" / "datos" / "log_sla_highest.md"

# Copia de trabajo (donde el pipeline realmente lee/escribe) -- se siembra desde el
# espejo al arrancar `ingest`. Ver CONTRATO DE ESCRITURA en el docstring de arriba.
STAGING_DIR = REPO / "wiki" / "1_proyectos" / "contexto_vivo" / "_staging_dashboard_delivery"
LOG_MD = STAGING_DIR / "log_performance_desarrollo.md"
LOG_COST_MD = STAGING_DIR / "log_costos_desarrollo.md"
LOG_SLA_MD = STAGING_DIR / "log_sla_highest.md"

TEMPLATE = SCRIPT_DIR.parent / "assets" / "dashboard_template.html"
DASHBOARD = REPO / "outputs" / "dashboard_performance_desarrollo.html"  # no es canon, directo


def seed_staging_from_mirror():
    """Siembra la copia de trabajo desde el espejo si todavia no existe. No pisa una
    copia ya sembrada en esta corrida -- si `ingest` corre dos veces sin pasar por
    /context_push, la segunda sigue mergeando sobre lo que la primera ya acumulo."""
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    for mirror, staged in (
        (MIRROR_LOG_MD, LOG_MD),
        (MIRROR_LOG_COST_MD, LOG_COST_MD),
        (MIRROR_LOG_SLA_MD, LOG_SLA_MD),
    ):
        if not staged.exists() and mirror.exists():
            shutil.copy2(mirror, staged)

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
MES_ALIAS = {"setiembre": "Septiembre"}
ESPACIOS_BASE = ["AD", "WS", "OB", "SER"]
ANIO_DEFAULT = 2026  # usado SOLO cuando el origen no trae columna de año (formato ticket-level mensual del PM).
# GOTCHA: cuando lleguen reportes mensuales de 2027, bump este valor (o mejor: pedirle
# al PM que incluya el año, y leerlo de ahí en vez de este default). Ver SKILL.md.

# --- Formato 1: ticket-level (export Jira mensual del PM) -------------------
REQUIRED = {
    "clave": ["clave de incidencia"],          # sin esto no hay espacio ni ticket -> ABORTA
    "mes":   ["mes"],                          # sin esto no hay eje X -> ABORTA
}
DEGRADABLE = {
    "tipo": ["tipo de incidencia"],            # falta -> "(sin tipo)" + warning
    "sp":   ["campo personalizado (story points)", "story points", "puntos de historia"],  # falta -> 0 + warning
    "epic": ["parent summary", "epic"],        # falta -> epic=None (sin dato) + warning
}
KNOWN_EXTRA = [
    "id de la incidencia", "resumen", "campo personalizado (id fintexa)", "prioridad",
    "estado", "versiones corregidas", "persona asignada", "id de la persona asignada",
    "principal", "clave principal", "fecha de vencimiento",
    "campo personalizado (fecha de inicio)",
]

# --- Formato 2: agregado por versión (backfills históricos) -----------------
VA_SIGNATURE = {"sp-us", "q-bugs"}  # si el header de una hoja contiene esto, es formato 2
VA_COLS = {
    "version":  ["version"],
    "producto": ["producto"],
    "anio":     ["ano publicacion"],
    "mes":      ["mes publicacion"],
    "sp_us":    ["sp-us"],
    "sp_bugs":  ["sp-bugs"],
    "q_us":     ["q-us"],
    "q_bugs":   ["q-bugs"],
    "sp_tot":   ["sp"],
    "q_tot":    ["q tickets"],
}
VA_PRODUCTO_MAP = {"wallet": "WS", "cobro": "AD"}  # ampliar acá si aparece OB/SER en un backfill futuro

# --- Formato 3: stock de horas mensual (factura de Fintexa) -----------------
STOCK_SHEET_RE = re.compile(r"stock de horas\s*-?\s*([a-z]{3})\s*(\d{2})")
MES3 = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}
STOCK_DEV_SECTIONS = {"operativo", "devops", "coe", "gestion", "kits", "deuda",
                      "onboarding", "onboarding legajos", "s.a.", "producto"}
STOCK_EXCL_SECTIONS = {"soporte", "it"}  # no son horas de desarrollo (decisión del usuario 2026-07-21)
STOCK_COMPONENTE_MAP = {
    "wallet": "WS", "wallet services": "WS", "wallet apk": "WS", "apk": "WS",
    "cobro": "AD", "admin / bp / bo": "AD",
    "onboarding": "OB", "onboarding pj": "OB", "ob-93 legajos": "OB",
    "deuda": "SER",
}  # ampliar acá si aparece un componente/producto nuevo en un stock futuro
STOCK_SPLIT = {"todos", "comite de arquitectura"}  # se reparte 50/50 WS-AD (decisión del usuario 2026-07-21)
STOCK_SKIP_PERIODS = {(2025, 1), (2025, 6)}
# GOTCHA: Ene25 y Jun25 quedan fuera de la ventana de SP (el log de delivery arranca en
# Jul25) y Ene25 además usa un layout viejo (columna "Rol" en vez de "Tipo", perfiles
# como "DEV Mob" que no existen en las tarifas posteriores). Se parsean igual para poder
# reportarlos en `inspect` a modo informativo, pero `ingest` los descarta (meta["skip"]).
STOCK_RATE_BLOCK_SKIP = {"horas soporte", "horas it"}  # bloques de tarifas que no son de desarrollo

# --- Formato 4: SLA de tickets Highest (export "tiempo por estado", CSV) -----
# Duraciones formato Jira "1M 2w 3d 4h 5m" (con signo opcional), TIEMPO CALENDARIO
# (no horas hábiles) — mismo parser que usa /dashboard_qa para el mismo tipo de export.
SLA_MONTH_MIN = 30.44 * 24 * 60
SLA_WEEK_MIN = 7 * 24 * 60
SLA_DAY_MIN = 24 * 60
SLA_DUR_RE = re.compile(r"(-?)(\d+)([Mwdhm])")

SLA_BACKLOG_COLS = ["Backlog", "Backlog-INF", "BACKLOG-10264"]
SLA_INPROGRESS_COLS = ["Asignado", "Bloqueado", "Con defecto", "En curso", "EN QA",
                        "In Progress-10299", "LISTO PARA DESARROLLO",
                        "SELECCIONADO PARA DESARROLLO-10268", "Selected for Development-10040"]
SLA_TERMINAL_COLS = ["Finalizada", "No aplica"]
# Firma que distingue este CSV de cualquier otro formato: columnas de tiempo-en-estado
# imprescindibles para el cálculo del SLA (no se adivina si faltan).
SLA_SIGNATURE_COLS = {"asignado", "backlog", "finalizada"}
SLA_META_COLS = {"Clave", "Tipo de Incidencia", "Resumen", "Estado", "Creada", "Prioridad",
                  "Story Points", "Versiones corregidas", "Resuelta"}


def parse_sla_duration(v):
    """'1M 2w 3d 4h 5m' (con signo opcional por token) -> minutos. '-'/vacío -> 0."""
    v = (v or "").strip()
    if v in ("", "-"):
        return 0.0
    total = 0.0
    for sign, n, unit in SLA_DUR_RE.findall(v):
        n = int(n) * (-1 if sign == "-" else 1)
        mult = {"M": SLA_MONTH_MIN, "w": SLA_WEEK_MIN, "d": SLA_DAY_MIN, "h": 60, "m": 1}[unit]
        total += n * mult
    return total


def parse_sla_dt(v):
    v = (v or "").strip()
    if not v:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return None


def norm(s):
    s = unicodedata.normalize("NFD", str(s or "").strip().lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def fnum(x):
    # Evita notación científica (%g la dispara a partir de ~1e6, y los totales de USD
    # del log de costos la cruzan fácil): formatea siempre en punto fijo.
    v = round(float(x), 2)
    if v == int(v):
        return str(int(v))
    return f"{v:.2f}".rstrip("0").rstrip(".")


def find_excels():
    if not RAW_DIR.exists():
        return []
    files = [p for pat in ("*.xlsx", "*.xlsm") for p in RAW_DIR.glob(pat)]
    return sorted(p for p in files if not p.name.startswith("~$"))


def find_csvs():
    if not RAW_DIR.exists():
        return []
    return sorted(p for p in RAW_DIR.glob("*.csv") if not p.name.startswith("~$"))


def sniff_csv_format(path):
    """Devuelve 'sla_highest' si el header matchea el formato conocido, None si no.
    No hay más de un formato CSV conocido hoy — si aparece uno nuevo, sumar su propia
    firma acá en vez de forzarlo dentro de esta."""
    with path.open(encoding="utf-8-sig", newline="") as f:
        header = next(csv.reader(f), [])
    h = {norm(x) for x in header if x}
    if SLA_SIGNATURE_COLS <= h and {"clave", "estado", "creada"} <= h:
        return "sla_highest"
    return None


# --- Lectura: formato 4 (SLA Highest, CSV) -----------------------------------
def read_sla_csv(path, today_dt):
    """Lee el export 'tiempo por estado' de tickets Highest y calcula, por ticket:
      - dias_finalizada: Creada -> entrada a 'Finalizada' (Creada + Σ tiempo en estados
        Backlog + en-progreso), solo si Estado == Finalizada.
      - dias_publicacion: Creada -> fecha REAL de release de la versión que lo resolvió
        (vía JSON compañero `<csv>.versions.json`, poblado por Claude consultando Jira
        por MCP antes de ingerir); si no hay versión released con fecha, degrada a la
        misma fecha proxy que dias_finalizada (fuente_publicacion="proxy_finalizada").
      - dias_abierto: para tickets NO Finalizada, días corridos desde Creada hasta
        `today_dt` (el reloj sigue corriendo — no se excluye de la métrica de riesgo,
        aunque sí se excluye de las medianas/promedios de "tiempo hasta resolución").
    """
    versions_path = path.parent / (path.stem + ".versions.json")
    jira_versions = {}
    has_versions_file = versions_path.exists()
    if has_versions_file:
        jira_versions = json.loads(versions_path.read_text(encoding="utf-8"))

    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    known = SLA_META_COLS | set(SLA_BACKLOG_COLS) | set(SLA_INPROGRESS_COLS) | set(SLA_TERMINAL_COLS)
    extras = [c for c in fieldnames if c not in known]

    tickets = {}
    n_rows, bad_dates, sin_jira_version = 0, [], []
    for row in rows:
        clave = (row.get("Clave") or "").strip()
        if not clave:
            continue
        creada = parse_sla_dt(row.get("Creada"))
        if creada is None:
            bad_dates.append(clave)
            continue
        n_rows += 1
        estado = (row.get("Estado") or "").strip()
        backlog_min = sum(parse_sla_duration(row.get(c)) for c in SLA_BACKLOG_COLS)
        inprog_min = sum(parse_sla_duration(row.get(c)) for c in SLA_INPROGRESS_COLS)
        entry_finalizada = creada + timedelta(minutes=backlog_min + inprog_min)

        dias_finalizada = dias_publicacion = fuente_publicacion = dias_abierto = None
        if estado == "Finalizada":
            dias_finalizada = round((entry_finalizada - creada).total_seconds() / 86400, 2)
            releases = jira_versions.get(clave) or []
            candidatas = [datetime.strptime(d, "%Y-%m-%d") for d in releases]
            if candidatas:
                posteriores = [d for d in candidatas if d >= creada]
                resolucion_dt = min(posteriores) if posteriores else max(candidatas)
                fuente_publicacion = "jira_release"
            else:
                resolucion_dt = entry_finalizada
                fuente_publicacion = "proxy_finalizada"
                sin_jira_version.append(clave)
            dias_publicacion = round((resolucion_dt - creada).total_seconds() / 86400, 2)
        else:
            dias_abierto = round((today_dt - creada).total_seconds() / 86400, 2)

        tickets[clave] = {
            "clave": clave,
            "espacio": clave.split("-")[0],
            "tipo": (row.get("Tipo de Incidencia") or "").strip(),
            "estado": estado,
            "creada": creada.strftime("%Y-%m-%d"),
            "anio": creada.year,
            "mes": MESES[creada.month - 1],
            "dias_finalizada": dias_finalizada,
            "dias_publicacion": dias_publicacion,
            "fuente_publicacion": fuente_publicacion,
            "dias_abierto": dias_abierto,
        }

    warnings = []
    if not has_versions_file:
        warnings.append(f"Sin JSON compañero '{versions_path.name}': todos los Finalizada de este lote degradan a fecha proxy (entrada a Finalizada) para 'días hasta publicación'. Generarlo consultando Jira (fixVersions/releaseDate) por MCP antes de ingerir.")
    elif sin_jira_version:
        warnings.append(f"{len(sin_jira_version)} ticket(s) Finalizada sin versión released con fecha en el JSON compañero — degradan a fecha proxy: {', '.join(sin_jira_version[:10])}" + (" …" if len(sin_jira_version) > 10 else ""))
    if bad_dates:
        warnings.append(f"{len(bad_dates)} ticket(s) con 'Creada' ilegible, EXCLUIDOS: {', '.join(bad_dates[:10])}" + (" …" if len(bad_dates) > 10 else ""))

    meta = {"file": path.name, "rows": n_rows, "extras": extras, "warnings": warnings,
            "format": "sla_highest", "kind": "sla"}
    return tickets, meta


# --- Log md: parseo y escritura (SLA Highest) --------------------------------
SLA_DETAIL_HEADER = ["Clave", "Espacio", "Tipo", "Estado", "Creada", "Días hasta Finalizada",
                      "Días hasta Publicación", "Fuente publicación", "Días abierto (a la ingesta)"]


def parse_sla_log():
    """Devuelve (tickets: {clave: dict}, lots) desde el log actual; ({}, []) si no existe.
    Igual que log_performance_qa.md: acumula UN registro por ticket (`Clave`), upsert
    por clave — un Highest abierto hoy puede seguir abierto (o ya finalizado) en el
    próximo export, y este log siempre refleja el dato más fresco de cada uno."""
    if not LOG_SLA_MD.exists():
        return {}, []
    text = LOG_SLA_MD.read_text(encoding="utf-8")
    tickets, lots = {}, []
    section = None
    for line in text.splitlines():
        if line.startswith("## "):
            if "Registro de lotes" in line:
                section = "lots"
            elif line.startswith("## Datos"):
                section = "data"
            else:
                section = None
            continue
        if not line.startswith("|") or line.startswith("|---") or section is None:
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if section == "lots" and len(cells) == 5 and cells[0] != "Fecha ingesta":
            lots.append(cells)
        elif section == "data" and len(cells) == 9 and cells[0] not in ("Clave",):
            (clave, espacio, tipo, estado, creada, dias_fin, dias_pub, fuente, dias_ab) = cells
            creada_dt = datetime.strptime(creada, "%Y-%m-%d")
            tickets[clave] = {
                "clave": clave, "espacio": espacio, "tipo": tipo, "estado": estado,
                "creada": creada, "anio": creada_dt.year, "mes": MESES[creada_dt.month - 1],
                "dias_finalizada": None if dias_fin == "—" else float(dias_fin),
                "dias_publicacion": None if dias_pub == "—" else float(dias_pub),
                "fuente_publicacion": None if fuente == "—" else fuente,
                "dias_abierto": None if dias_ab == "—" else float(dias_ab),
            }
    return tickets, lots


def sla_coverage_str(tickets):
    if not tickets:
        return "sin datos"
    meses = sorted({t["creada"][:7] for t in tickets})
    rango = meses[0] if meses[0] == meses[-1] else f"{meses[0]} – {meses[-1]}"
    espacios = sorted({t["espacio"] for t in tickets})
    return f"{rango}, {' + '.join(espacios)}"


def build_sla_resumen(tickets):
    """Tabla de sanity check por mes de CREACIÓN: no alimenta el dashboard (que
    recalcula todo del detalle en JS), es para que un humano audite de un vistazo."""
    by_mes = defaultdict(list)
    for t in tickets:
        by_mes[t["creada"][:7]].append(t)
    out = ["| Mes (creación) | Creados | Finalizados | Abiertos | Mediana días hasta Finalizada | Mediana días hasta Publicación | Máx días abierto |",
           "|---|---|---|---|---|---|---|"]
    for mes in sorted(by_mes):
        rows = by_mes[mes]
        finalizados = [t for t in rows if t["estado"] == "Finalizada"]
        abiertos = [t for t in rows if t["estado"] != "Finalizada"]
        dias_fin = [t["dias_finalizada"] for t in finalizados if t["dias_finalizada"] is not None]
        dias_pub = [t["dias_publicacion"] for t in finalizados if t["dias_publicacion"] is not None]
        max_abierto = max([t["dias_abierto"] for t in abiertos if t["dias_abierto"] is not None], default=None)
        m_fin = f"{statistics.median(dias_fin):.1f}" if dias_fin else "—"
        m_pub = f"{statistics.median(dias_pub):.1f}" if dias_pub else "—"
        m_ab = f"{max_abierto:.1f}" if max_abierto is not None else "—"
        out.append(f"| {mes} | {len(rows)} | {len(finalizados)} | {len(abiertos)} | {m_fin} | {m_pub} | {m_ab} |")
    return out


def write_sla_log(tickets_by_clave, lots, today):
    tickets = sorted(tickets_by_clave.values(), key=lambda t: (t["creada"], t["clave"]))
    resumen = build_sla_resumen(tickets)
    detalle = ["| " + " | ".join(SLA_DETAIL_HEADER) + " |", "|" + "---|" * len(SLA_DETAIL_HEADER)]
    for t in tickets:
        detalle.append(
            f"| {t['clave']} | {t['espacio']} | {t['tipo']} | {t['estado']} | {t['creada']} | "
            f"{fnum(t['dias_finalizada']) if t['dias_finalizada'] is not None else '—'} | "
            f"{fnum(t['dias_publicacion']) if t['dias_publicacion'] is not None else '—'} | "
            f"{t['fuente_publicacion'] or '—'} | "
            f"{fnum(t['dias_abierto']) if t['dias_abierto'] is not None else '—'} |"
        )
    lots_tbl = ["| Fecha ingesta | Archivo fuente | Cobertura | Tickets tocados | Destino histórico |",
                "|---|---|---|---|---|"] + ["| " + " | ".join(l) + " |" for l in lots]
    last_lot = lots[-1] if lots else ["—"] * 5

    finalizados = [t for t in tickets if t["estado"] == "Finalizada"]
    abiertos = [t for t in tickets if t["estado"] != "Finalizada"]
    con_proxy = sum(1 for t in finalizados if t["fuente_publicacion"] == "proxy_finalizada")
    dias_fin_all = [t["dias_finalizada"] for t in finalizados if t["dias_finalizada"] is not None]
    mediana_fin_global = fnum(statistics.median(dias_fin_all)) if dias_fin_all else "—"
    peor_abierto = max(abiertos, key=lambda t: t["dias_abierto"] or 0, default=None)

    md = f"""# Log de SLA de tickets Highest — Base de datos de la métrica "tiempo de resolución de urgencias"

> **Última ingesta:** {today} — {last_lot[1]} ({last_lot[2]}, {last_lot[3]}).
>
> Este archivo es la **base de datos acumulada**, por ticket, de la métrica de SLA de tickets Prioridad=Highest (reclamos de clientes, bugs, incendios) del dashboard [`outputs/dashboard_performance_desarrollo.html`](../../outputs/dashboard_performance_desarrollo.html), mantenida por la skill [`/dashboard_delivery`](../../.claude/skills/dashboard_delivery/SKILL.md). El usuario deja en `raw/` un export de Jira "tiempo por estado" filtrado a Prioridad=Highest (una fila por ticket, tiempo acumulado por estado — mismo formato que usa `/dashboard_qa`). Cada ingesta hace **upsert por `Clave`**: el ticket que reaparece se actualiza con el dato fresco, el que no aparece en el export nuevo se conserva tal cual. Los agregados se recalculan enteros desde este detalle en cada corrida.
>
> ⚠️ **Requiere un paso manual de Claude antes de ingerir:** este export no trae fecha de release de versión, así que antes de correr `pipeline.py ingest` hay que consultar Jira por MCP (`fixVersions`/`releaseDate`) para cada ticket Finalizada del CSV y volcarlo a `<mismo nombre>.versions.json` en `raw/` (mapa Clave -> lista de fechas ISO de versiones released). Sin ese archivo, "días hasta Publicación" degrada entero a la fecha proxy de "días hasta Finalizada" — ver metodología abajo.

## Metodología / criterios de agregación

- **Universo:** tickets Prioridad=Highest del export (reclamos, bugs urgentes, incendios) — no todos los tickets del backlog.
- **Días hasta Finalizada** = `Creada` + Σ(tiempo en Backlog* + Asignado + Bloqueado + Con defecto + En curso + EN QA + Listo/Seleccionado para desarrollo), es decir, todo el tiempo del export EXCEPTO el que corre dentro de "Finalizada"/"No aplica" (estados terminales cuyo reloj sigue corriendo hasta el export, no se congela). Solo tiene valor para tickets con Estado = Finalizada — el reloj de los demás no paró.
- **Días hasta Publicación** = `Creada` hasta la fecha REAL de release de la versión que resolvió el ticket (`fixVersions`/`releaseDate` de Jira, vía el JSON compañero). Si el ticket no tiene ninguna versión released con fecha en ese JSON, degrada a la misma fecha proxy que "Días hasta Finalizada" (`fuente_publicacion = proxy_finalizada`) — hoy {con_proxy} de {len(finalizados)} finalizados están en ese caso.
- **Días abierto (a la ingesta):** para tickets que NO llegaron a Finalizada, días corridos entre `Creada` y el momento de esta ingesta — el reloj **sigue corriendo**: estos tickets NO se excluyen de la vista de riesgo (aparecen igual, con su conteo de días creciendo), solo se excluyen de las medianas/promedios de "tiempo hasta resolución" (no tiene sentido promediar una medición que no terminó). Valor congelado a la fecha de esta ingesta — no se recalcula solo, hace falta una ingesta nueva para actualizarlo.
- **Mediana global (Días hasta Finalizada), {len(finalizados)} tickets resueltos:** {mediana_fin_global} días.
- **Peor caso abierto ahora mismo:** {f"{peor_abierto['clave']} — {fnum(peor_abierto['dias_abierto'])} días corridos desde {peor_abierto['creada']} (Estado: {peor_abierto['estado']})" if peor_abierto else "—"}.
- **Eje X del dashboard = mes de `Creada`** (decisión del usuario, 2026-07-27): mide "las urgencias creadas en el mes X tardaron Y días en resolverse", no "cuánto se resolvió ese mes" — los meses recientes con tickets aún abiertos quedan marcados como cohorte incompleta (en itálica en el eje) porque su mediana todavía no incluye a los lentos que siguen corriendo.

## Registro de lotes ingeridos

{chr(10).join(lots_tbl)}

## Resumen mensual (sanity check — el dashboard recalcula todo del detalle)

{chr(10).join(resumen)}

## Datos — detalle por ticket

{chr(10).join(detalle)}
"""
    LOG_SLA_MD.write_text(md, encoding="utf-8")


# --- Lectura: formato 1 (ticket-level) ---------------------------------------
def map_headers(header_row):
    """Devuelve (colmap, extras, missing_required, missing_degradable)."""
    headers = {norm(h): i for i, h in enumerate(header_row) if h not in (None, "")}
    colmap, missing_req, missing_deg = {}, [], []
    for field, aliases in {**REQUIRED, **DEGRADABLE}.items():
        idx = next((headers[a] for a in aliases if a in headers), None)
        if idx is not None:
            colmap[field] = idx
        elif field in REQUIRED:
            missing_req.append(field)
        else:
            missing_deg.append(field)
    known = {a for al in list(REQUIRED.values()) + list(DEGRADABLE.values()) for a in al} | set(KNOWN_EXTRA)
    extras = [str(header_row[i]) for h, i in headers.items() if h not in known]
    return colmap, extras, missing_req, missing_deg


def read_ticket_ws(ws, colmap, extras, missing_deg, path):
    warnings = [f"columna '{f}' ausente — se degrada a valor por defecto" for f in missing_deg]
    agg = defaultdict(lambda: {"tickets": 0, "sp": 0.0})
    bad_meses, sin_sp, n_rows = set(), [], 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        clave = row[colmap["clave"]] if colmap["clave"] < len(row) else None
        if clave in (None, ""):
            continue
        clave = str(clave).strip()
        mes_raw = row[colmap["mes"]] if colmap["mes"] < len(row) else None
        mes = MES_ALIAS.get(norm(mes_raw), str(mes_raw or "").strip().capitalize())
        if mes not in MESES:
            bad_meses.add(repr(mes_raw))
            continue
        def cell(field, default):
            i = colmap.get(field)
            v = row[i] if i is not None and i < len(row) else None
            return default if v in (None, "") else v
        tipo = str(cell("tipo", "(sin tipo)")).strip()
        epic_raw = cell("epic", None)
        epic = str(epic_raw).strip() if epic_raw is not None else None
        sp_v = cell("sp", None)
        if sp_v is None:
            sin_sp.append(clave)
            sp = 0.0
        else:
            try:
                sp = float(sp_v)
            except (TypeError, ValueError):
                sin_sp.append(clave)
                sp = 0.0
        espacio = clave.split("-")[0]
        key = (ANIO_DEFAULT, mes, espacio, tipo, epic)
        agg[key]["tickets"] += 1
        agg[key]["sp"] += sp
        n_rows += 1

    if bad_meses:
        warnings.append(f"{len(bad_meses)} valor(es) de Mes no reconocidos y EXCLUIDOS: {sorted(bad_meses)}")
    if sin_sp:
        warnings.append(f"{len(sin_sp)} tickets sin SP (computan 0): {', '.join(sin_sp[:10])}"
                        + (" …" if len(sin_sp) > 10 else ""))
    if "epic" in missing_deg:
        warnings.append("Sin columna de Epic: estos meses no tendrán desglose por Epic en el dashboard (esperado, no se inventa un bucket).")
    records = [{"anio": a, "mes": m, "espacio": e, "tipo": t, "epic": ep,
                "tickets": v["tickets"], "sp": round(v["sp"], 2)}
               for (a, m, e, t, ep), v in agg.items()]
    meta = {"file": path.name, "sheet": ws.title, "rows": n_rows, "extras": extras, "warnings": warnings,
            "format": "ticket-level", "kind": "delivery"}
    return records, meta


# --- Lectura: formato 2 (agregado por versión, backfills) -------------------
def read_version_agg_ws(ws, path):
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    idx = {norm(h): i for i, h in enumerate(row1) if h not in (None, "")}

    def col(field):
        for a in VA_COLS[field]:
            if a in idx:
                return idx[a]
        return None

    c = {f: col(f) for f in VA_COLS}
    agg = defaultdict(lambda: {"tickets": 0, "sp": 0.0})
    unknown_prod, n_rows, mismatches, bad_periods = set(), 0, 0, set()

    def get(row, field):
        i = c[field]
        return row[i] if i is not None and i < len(row) else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if get(row, "version") in (None, ""):
            continue
        prod_raw = get(row, "producto")
        espacio = VA_PRODUCTO_MAP.get(norm(prod_raw))
        if espacio is None:
            unknown_prod.add(str(prod_raw))
            continue
        try:
            anio = int(get(row, "anio"))
            mes_num = int(get(row, "mes"))
        except (TypeError, ValueError):
            bad_periods.add((get(row, "anio"), get(row, "mes")))
            continue
        if not (1 <= mes_num <= 12):
            bad_periods.add((anio, mes_num))
            continue
        mes = MESES[mes_num - 1]
        spus = float(get(row, "sp_us") or 0)
        spbugs = float(get(row, "sp_bugs") or 0)
        qus = int(get(row, "q_us") or 0)
        qbugs = int(get(row, "q_bugs") or 0)
        sp_tot, q_tot = get(row, "sp_tot"), get(row, "q_tot")
        if sp_tot is not None and abs((spus + spbugs) - float(sp_tot)) > 0.01:
            mismatches += 1
        if q_tot is not None and (qus + qbugs) != int(q_tot):
            mismatches += 1
        n_rows += 1
        if qus or spus:
            k = (anio, mes, espacio, "Historia", None)
            agg[k]["tickets"] += qus
            agg[k]["sp"] += spus
        if qbugs or spbugs:
            k = (anio, mes, espacio, "Error", None)
            agg[k]["tickets"] += qbugs
            agg[k]["sp"] += spbugs

    warnings = ["Formato agregado por versión, SIN Epic: estos meses no tendrán desglose por Epic en el dashboard (esperado)."]
    if unknown_prod:
        warnings.append(f"PRODUCTO no reconocido, fila(s) descartada(s) — mapear en VA_PRODUCTO_MAP si es un espacio válido: {sorted(unknown_prod)}")
    if bad_periods:
        warnings.append(f"AÑO/MES inválido, fila(s) descartada(s): {sorted(str(x) for x in bad_periods)}")
    if mismatches:
        warnings.append(f"{mismatches} fila(s) con SP o Q total inconsistente respecto de US+BUGS (se usó igual la suma US+BUGS)")

    records = [{"anio": a, "mes": m, "espacio": e, "tipo": t, "epic": ep,
                "tickets": v["tickets"], "sp": round(v["sp"], 2)}
               for (a, m, e, t, ep), v in agg.items()]
    meta = {"file": path.name, "sheet": ws.title, "rows": n_rows, "extras": [], "warnings": warnings,
            "format": "agregado-por-version", "kind": "delivery"}
    return records, meta


# --- Lectura: formato 3 (stock de horas, factura Fintexa) -------------------
def _padded_grid(ws):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    w = max((len(r) for r in rows), default=0)
    return [r + [None] * (w - len(r)) for r in rows], w


def _find_stock_header(grid):
    """Busca, en las primeras 6 filas, la fila con la columna 'Recursos' (tolera que
    el proveedor agregue una columna PRODUCTO a la izquierda o reordene el resto)."""
    for i, row in enumerate(grid[:6]):
        for j, c in enumerate(row):
            if norm(c) == "recursos":
                return i, j
    return None, None


def read_stock_rates(grid, w):
    """Extrae la tabla de precios del Excel, en cualquiera de sus dos formas conocidas:
    simple (Perfil | Valor Hora (USD)) o por bloques (Tipo | Suma de Horas Mes | Valor
    Hora), ignorando los bloques de Soporte/IT (no son tarifas de desarrollo)."""
    rates = {}
    for i, row in enumerate(grid):
        for j, c in enumerate(row):
            n = norm(c)
            if n == "perfil" and j + 1 < w and "valor hora" in norm(row[j + 1]):
                k = i + 1
                while k < len(grid) and grid[k][j] not in (None, ""):
                    if isinstance(grid[k][j + 1], (int, float)):
                        rates.setdefault(norm(grid[k][j]), float(grid[k][j + 1]))
                    k += 1
            elif n == "tipo" and j + 2 < w and "valor hora" in norm(row[j + 2]):
                label = next((norm(grid[k][j]) for k in range(i - 1, max(-1, i - 4), -1)
                              if grid[k][j] and norm(grid[k][j]).startswith("horas")), "")
                if label in STOCK_RATE_BLOCK_SKIP:
                    continue
                k = i + 1
                while k < len(grid) and grid[k][j] not in (None, ""):
                    nm = norm(grid[k][j])
                    if nm.startswith("total"):
                        break
                    if isinstance(grid[k][j + 2], (int, float)):
                        rates.setdefault(nm, float(grid[k][j + 2]))
                    k += 1
    return rates


def read_stock_ws(ws, path, anio, mes_num):
    """Devuelve (rows, meta). `rows` son horas de DESARROLLO agregadas por (espacio,
    perfil) — sin costo todavía: el costo depende de la tabla de tarifas resuelta a
    nivel de lote en cmd_inspect/cmd_ingest (algunos meses no traen tarifa propia y
    hay que heredarla del mes conocido más cercano, algo que no se puede resolver
    hoja por hoja)."""
    grid, w = _padded_grid(ws)
    i0, jr = _find_stock_header(grid)
    idx = {norm(c): j for j, c in enumerate(grid[i0]) if c not in (None, "")}
    ctipo = idx.get("tipo", idx.get("rol"))
    ccomp = idx.get("componente / proyecto")
    chm = idx.get("horas mes")
    own_rates = read_stock_rates(grid, w)

    sect = None
    horas = defaultdict(lambda: defaultdict(float))  # espacio -> tipo -> horas
    unknown_comp = set()
    for row in grid[i0:]:
        if any(norm(c).startswith("total stock") for c in row):
            break  # corta antes de bloques retroactivos post-total (caso SEP25)
        a = norm(row[jr - 1]) if jr - 1 >= 0 else ""
        b = row[jr]
        nb = norm(b)
        if nb == "recursos":
            sect = a if a in STOCK_DEV_SECTIONS | STOCK_EXCL_SECTIONS else None
            continue
        if a in STOCK_DEV_SECTIONS | STOCK_EXCL_SECTIONS:
            sect = a
        if b in (None, "") or nb.startswith(("subtotal", "total")):
            continue
        hm = row[chm] if chm is not None else None
        if not isinstance(hm, (int, float)) or sect in STOCK_EXCL_SECTIONS:
            continue
        tipo = norm(row[ctipo]) if ctipo is not None else ""
        comp = norm(row[ccomp]) if ccomp is not None else ""
        if comp in STOCK_COMPONENTE_MAP:
            horas[STOCK_COMPONENTE_MAP[comp]][tipo] += float(hm)
        elif comp in STOCK_SPLIT:
            horas["WS"][tipo] += float(hm) / 2
            horas["AD"][tipo] += float(hm) / 2
        else:
            unknown_comp.add(str(row[ccomp]))

    rows = [{"anio": anio, "mes_num": mes_num, "espacio": espacio, "tipo": tipo, "horas": round(hs, 2)}
            for espacio, por_tipo in horas.items() for tipo, hs in por_tipo.items() if hs]

    skip = (anio, mes_num) in STOCK_SKIP_PERIODS
    warnings = []
    if unknown_comp:
        warnings.append(f"Componente no reconocido, fila(s) descartada(s) — mapear en STOCK_COMPONENTE_MAP/STOCK_SPLIT: {sorted(unknown_comp)}")
    if not own_rates:
        warnings.append("Sin tabla de precios en esta hoja: se hereda la tarifa del mes conocido más cercano.")
    if skip:
        warnings.append("Fuera de la ventana de SP publicados — NO se ingiere (se muestra solo a modo informativo).")
    meta = {"file": path.name, "sheet": ws.title, "rows": sum(len(v) for v in horas.values()),
            "extras": [], "warnings": warnings, "format": "stock-horas", "kind": "cost",
            "anio": anio, "mes_num": mes_num, "own_rates": own_rates, "skip": skip}
    return rows, meta


# --- Dispatcher ---------------------------------------------------------------
def sniff_and_read(path):
    """Detecta el/los formato(s) presentes en el Excel por header y los parsea.
    Un mismo workbook puede traer varias hojas de stock de horas (ej. un archivo de
    control con varios meses); por eso devuelve una LISTA de (records, meta), no un
    único resultado. Aborta con mensaje claro si ninguna hoja matchea un formato
    conocido — no se adivina el mapeo."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results = []
    va_done, ticket_done = False, False
    for ws in wb.worksheets:
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        h = {norm(x) for x in row1 if x not in (None, "")}
        if not va_done and VA_SIGNATURE <= h:
            results.append(read_version_agg_ws(ws, path))
            va_done = True
            continue
        if not ticket_done:
            colmap, extras, missing_req, missing_deg = map_headers(row1)
            if not missing_req:
                results.append(read_ticket_ws(ws, colmap, extras, missing_deg, path))
                ticket_done = True
                continue
        m = STOCK_SHEET_RE.search(norm(ws.title))
        if m:
            grid, w = _padded_grid(ws)
            i0, jr = _find_stock_header(grid)
            if i0 is not None:
                idx = {norm(c) for c in grid[i0] if c not in (None, "")}
                if {"componente / proyecto", "horas mes"} <= idx:
                    anio, mes_num = 2000 + int(m.group(2)), MES3[m.group(1)]
                    results.append(read_stock_ws(ws, path, anio, mes_num))
    if not results:
        sys.exit(f"[ABORT] {path.name}: ninguna hoja coincide con un formato conocido "
                 f"(ni ticket-level de Jira con columna 'Clave de incidencia'+'Mes', ni agregado-por-versión "
                 f"con columnas SP-US/Q-BUGS, ni stock de horas con 'Componente / Proyecto'+'Horas Mes'). "
                 f"Hojas: {wb.sheetnames}. Revisar el archivo con el usuario antes de reintentar — no adivinar el mapeo.")
    return results


# --- Log md: parseo y escritura (delivery) --------------------------------------
def parse_log():
    """Devuelve (records, lots) desde el log actual; ([], []) si no existe (bootstrap).
    Compatible con el formato legacy sin columna Año (6 celdas, se asume ANIO_DEFAULT)."""
    if not LOG_MD.exists():
        return [], []
    text = LOG_MD.read_text(encoding="utf-8")
    records, lots = [], []
    section = None
    for line in text.splitlines():
        if line.startswith("## "):
            if "Registro de lotes" in line:
                section = "lots"
            elif line.startswith("## Datos"):
                section = "data"
            else:
                section = None
            continue
        if not line.startswith("|") or line.startswith("|---") or section is None:
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if section == "lots" and len(cells) == 6 and cells[0] != "Fecha ingesta":
            lots.append(cells)
        elif section == "data" and len(cells) == 7 and cells[1] in MESES:
            anio, mes, espacio, tipo, epic, tickets, sp = cells
            records.append({"anio": int(anio), "mes": mes, "espacio": espacio, "tipo": tipo,
                            "epic": None if epic == "—" else epic, "tickets": int(tickets), "sp": float(sp)})
        elif section == "data" and len(cells) == 6 and cells[0] in MESES:
            mes, espacio, tipo, epic, tickets, sp = cells
            records.append({"anio": ANIO_DEFAULT, "mes": mes, "espacio": espacio, "tipo": tipo,
                            "epic": None if epic == "—" else epic, "tickets": int(tickets), "sp": float(sp)})
    return records, lots


def sort_records(records):
    records.sort(key=lambda r: (r["anio"], MESES.index(r["mes"]), r["espacio"], r["tipo"], r["epic"] or ""))
    return records


def period_label(anio, mes):
    return f"{mes} {anio}"


def coverage_str(records):
    if not records:
        return "sin datos"
    points = sorted({(r["anio"], MESES.index(r["mes"])) for r in records})
    lo, hi = points[0], points[-1]
    espacios = sorted({r["espacio"] for r in records})
    lo_lbl = period_label(lo[0], MESES[lo[1]])
    hi_lbl = period_label(hi[0], MESES[hi[1]])
    rango = lo_lbl if lo == hi else f"{lo_lbl} – {hi_lbl}"
    return f"{rango}, {' + '.join(espacios)}"


def write_log(records, lots, today):
    resumen = build_resumen(records)
    detalle = ["| Año | Mes | Espacio | Tipo | Epic | Tickets | SP |", "|---|---|---|---|---|---|---|"]
    for r in records:
        epic_cell = r["epic"] if r["epic"] else "—"
        detalle.append(f"| {r['anio']} | {r['mes']} | {r['espacio']} | {r['tipo']} | {epic_cell} | {r['tickets']} | {fnum(r['sp'])} |")
    lots_tbl = ["| Fecha ingesta | Archivo fuente | Cobertura | Tickets | SP | Destino histórico |",
                "|---|---|---|---|---|---|"] + ["| " + " | ".join(l) + " |" for l in lots]
    last_lot = lots[-1] if lots else ["—"] * 6
    md = f"""# Log de Performance de Desarrollo — Base de datos del dashboard "Performance de desarrollo"

> **Última ingesta:** {today} — {last_lot[1]} ({last_lot[2]}, {last_lot[3]} tickets).
>
> Este archivo es la **base de datos acumulada** del dashboard [`outputs/dashboard_performance_desarrollo.html`](../../outputs/dashboard_performance_desarrollo.html), mantenida por la skill [`/dashboard_delivery`](../../.claude/skills/dashboard_delivery/SKILL.md). El PM deja todos los principios de mes un Excel en `raw/` con lo **publicado en producción** (tickets Historia/Error con versión corregida); ocasionalmente se suman backfills históricos puntuales de otras fuentes. Cada ingesta: (1) el pipeline mergea acá las filas nuevas con granularidad año × mes × espacio × tipo × epic — el Excel nuevo PISA los combos año×mes×espacio que trae —, (2) se regenera el dashboard embebiendo esta tabla como JSON, (3) el Excel rota a `4_archivos/historial_raw/`. **No hace falta releer los Excel históricos: este log es la fuente.**
>
> ⚠️ **Este reporte NO alimenta el conocimiento de producto de la wiki** (indicación del usuario 2026-07-13): es una métrica de management para medir al equipo de desarrollo por lo ENTREGADO en producción. El conocimiento de producto de las publicaciones lo maneja `/sync_releases`. El costo de ese desarrollo (USD/SP) se mide aparte, en [`log_costos_desarrollo.md`](log_costos_desarrollo.md).

## Metodología / criterios de agregación

- **Fuente:** reporte mensual del PM (export de Jira, formato ticket-level) y, puntualmente, backfills históricos de otras fuentes (formato agregado por versión). Se cuenta todo ticket/versión listado, incluidos los tickets en estado "No aplica" (decisión del usuario 2026-07-13: si está en el reporte de publicaciones, cuenta como entregado).
- **Espacio:** prefijo de la clave del ticket (WS-123 → WS) en el formato ticket-level; columna PRODUCTO (WALLET→WS, COBRO→AD) en el formato agregado por versión.
- **Epic:** columna "Parent summary" de Jira (formato ticket-level), con trim de espacios. El formato agregado por versión **no trae Epic** — esos registros quedan con Epic vacío (`—`) y **no aparecen en las vistas "por Epic" del dashboard** (no se agrupan en un bucket "sin epic": esos meses simplemente no muestran datos en esa métrica, por pedido explícito del usuario 2026-07-14).
- **SP nulos → 0**; tickets sin clave se descartan; valores de "Mes"/"Año" no reconocidos se excluyen y se reportan.
- **Año:** el formato ticket-level mensual del PM no trae columna de año — se asume `ANIO_DEFAULT` (ver `pipeline.py`, hoy 2026; hay que bumpearlo a mano cuando lleguen reportes de 2027). El formato agregado por versión sí trae "AÑO PUBLICACIÓN" explícito.
- **Epics BAU fijas** (colores constantes en el dashboard): SOPORTE (rojo), REGRESIONES WS, REGRESIONES AD, REGRESIONES OB, REGRESIONES SER, COE, INICIATIVAS TECNICAS. El resto de epics se pinta en escala de grises.

## Registro de lotes ingeridos

{chr(10).join(lots_tbl)}

## Resumen mensual (tickets / SP publicados)

{chr(10).join(resumen)}

## Datos — detalle año × mes × espacio × tipo × epic

{chr(10).join(detalle)}
"""
    LOG_MD.write_text(md, encoding="utf-8")


def build_resumen(records):
    ctrl = defaultdict(lambda: [0, 0.0])
    tot = defaultdict(lambda: [0, 0.0])
    espacios = sorted({r["espacio"] for r in records})
    for r in records:
        ctrl[(r["anio"], r["mes"], r["espacio"])][0] += r["tickets"]
        ctrl[(r["anio"], r["mes"], r["espacio"])][1] += r["sp"]
        tot[r["espacio"]][0] += r["tickets"]
        tot[r["espacio"]][1] += r["sp"]
    periods = sorted({(r["anio"], r["mes"]) for r in records}, key=lambda p: (p[0], MESES.index(p[1])))
    out = ["| Año | Mes | " + " | ".join(espacios) + " | Total |", "|---|---|" + "---|" * (len(espacios) + 1)]
    for (anio, mes) in periods:
        cells, t_tk, t_sp = [], 0, 0.0
        for e in espacios:
            tk, sp = ctrl.get((anio, mes, e), [0, 0.0])
            t_tk += tk
            t_sp += sp
            cells.append(f"{tk} tk / {fnum(sp)} SP" if tk else "—")
        out.append(f"| {anio} | {mes} | {' | '.join(cells)} | **{t_tk} tk / {fnum(t_sp)} SP** |")
    cells = [f"**{tot[e][0]} tk / {fnum(tot[e][1])} SP**" for e in espacios]
    g_tk = sum(tot[e][0] for e in espacios)
    g_sp = sum(tot[e][1] for e in espacios)
    out.append(f"| **Total** | **histórico** | {' | '.join(cells)} | **{g_tk} tk / {fnum(g_sp)} SP** |")
    return out


# --- Log md: parseo y escritura (costos) ------------------------------------
def parse_cost_log():
    """Devuelve (records, lots, rate_rows) desde el log de costos actual; ([], [], [])
    si no existe (bootstrap). `rate_rows` es el registro de tarifas por perfil que
    permite el carry-forward sin releer Excels históricos."""
    if not LOG_COST_MD.exists():
        return [], [], []
    text = LOG_COST_MD.read_text(encoding="utf-8")
    records, lots, rate_rows = [], [], []
    section = None
    for line in text.splitlines():
        if line.startswith("## "):
            if "Registro de lotes" in line:
                section = "lots"
            elif line.startswith("## Registro de tarifas"):
                section = "rates"
            elif line.startswith("## Datos"):
                section = "data"
            else:
                section = None
            continue
        if not line.startswith("|") or line.startswith("|---") or section is None:
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if section == "lots" and len(cells) == 6 and cells[0] != "Fecha ingesta":
            lots.append(cells)
        elif section == "rates" and len(cells) == 5 and cells[0].isdigit():
            anio, mes, perfil, usd_hora, origen = cells
            rate_rows.append({"anio": int(anio), "mes": mes, "mes_num": MESES.index(mes) + 1,
                              "perfil": perfil, "usd_hora": float(usd_hora), "origen": origen})
        elif section == "data" and len(cells) == 5 and cells[0].isdigit():
            anio, mes, espacio, horas, usd = cells
            records.append({"anio": int(anio), "mes": mes, "espacio": espacio,
                            "horas": float(horas), "usd": float(usd)})
    return records, lots, rate_rows


def sort_cost_records(records):
    records.sort(key=lambda r: (r["anio"], MESES.index(r["mes"]), r["espacio"]))
    return records


def build_cost_resumen(records):
    ctrl = defaultdict(lambda: [0.0, 0.0])
    tot = defaultdict(lambda: [0.0, 0.0])
    espacios = sorted({r["espacio"] for r in records})
    for r in records:
        ctrl[(r["anio"], r["mes"], r["espacio"])][0] += r["horas"]
        ctrl[(r["anio"], r["mes"], r["espacio"])][1] += r["usd"]
        tot[r["espacio"]][0] += r["horas"]
        tot[r["espacio"]][1] += r["usd"]
    periods = sorted({(r["anio"], r["mes"]) for r in records}, key=lambda p: (p[0], MESES.index(p[1])))
    out = ["| Año | Mes | " + " | ".join(espacios) + " | Total |", "|---|---|" + "---|" * (len(espacios) + 1)]
    for (anio, mes) in periods:
        cells, t_h, t_u = [], 0.0, 0.0
        for e in espacios:
            h, u = ctrl.get((anio, mes, e), [0.0, 0.0])
            t_h += h
            t_u += u
            cells.append(f"{fnum(h)} hs / ${fnum(u)}" if h else "—")
        out.append(f"| {anio} | {mes} | {' | '.join(cells)} | **{fnum(t_h)} hs / ${fnum(t_u)}** |")
    cells = [f"**{fnum(tot[e][0])} hs / ${fnum(tot[e][1])}**" for e in espacios]
    g_h = sum(tot[e][0] for e in espacios)
    g_u = sum(tot[e][1] for e in espacios)
    out.append(f"| **Total** | **histórico** | {' | '.join(cells)} | **{fnum(g_h)} hs / ${fnum(g_u)}** |")
    return out


def write_cost_log(records, lots, rate_rows, today):
    resumen = build_cost_resumen(records)
    detalle = ["| Año | Mes | Espacio | Horas | USD |", "|---|---|---|---|---|"]
    for r in records:
        detalle.append(f"| {r['anio']} | {r['mes']} | {r['espacio']} | {fnum(r['horas'])} | {fnum(r['usd'])} |")
    rate_rows_sorted = sorted(rate_rows, key=lambda r: (r["anio"], r["mes_num"], r["perfil"]))
    tarifas = ["| Año | Mes | Perfil | USD/h | Origen |", "|---|---|---|---|---|"]
    for r in rate_rows_sorted:
        tarifas.append(f"| {r['anio']} | {r['mes']} | {r['perfil']} | {fnum(r['usd_hora'])} | {r['origen']} |")
    lots_tbl = ["| Fecha ingesta | Archivo fuente | Cobertura | Horas | USD | Destino histórico |",
                "|---|---|---|---|---|---|"] + ["| " + " | ".join(l) + " |" for l in lots]
    last_lot = lots[-1] if lots else ["—"] * 6
    md = f"""# Log de Costos de Desarrollo — Base de datos de la métrica USD/SP

> **Última ingesta:** {today} — {last_lot[1]} ({last_lot[2]}, {last_lot[3]} hs / ${last_lot[4]}).
>
> Este archivo es la **base de datos acumulada** de costo de desarrollo, mantenida por la skill [`/dashboard_delivery`](../../.claude/skills/dashboard_delivery/SKILL.md) a partir del **stock de horas mensual que factura Fintexa** (Excel `Stock de Horas - <Mes><Año>` depositado en `raw/`). Junto con [`log_performance_desarrollo.md`](log_performance_desarrollo.md) (SP publicados) alimenta la pestaña **USD por SP** del dashboard [`outputs/dashboard_performance_desarrollo.html`](../../outputs/dashboard_performance_desarrollo.html): cuánto cuesta un punto de historia y cómo evoluciona ese costo mes a mes.

## Metodología / criterios de agregación

- **Horas de desarrollo:** todas las secciones del stock salvo `SOPORTE` e `IT` (no son desarrollo de producto). Incluye `OPERATIVO`, `DEVOPS`, `COE`, `GESTION`, `KITS`/`DEUDA`, `ONBOARDING`/`S.A.`.
- **Cálculo del costo — fila por fila, nunca promediado:** `USD = Horas Mes × Valor Hora del perfil de esa fila`. El costo total de un espacio es la suma de ese producto en todas sus filas; el "USD/hora promedio" que resulta de dividir el total es un efecto de la mezcla de perfiles asignados ese mes, no un insumo del cálculo.
- **Imputación por Componente / Proyecto:** Wallet / Wallet Services / Wallet APK → **WS**; Cobro y Admin / BP / BO → **AD**; Onboarding / Onboarding PJ / OB-93 Legajos → **OB**; Deuda → **SER**; Todos y Comité de Arquitectura → **50% WS / 50% AD** (esfuerzo transversal). Componente no reconocido → fila descartada y advertida, nunca imputada por adivinanza.
- **Tarifas faltantes (carry-forward):** cuando el Excel de un mes no trae su propia tabla de precios, se hereda la tarifa del mes conocido más cercano (si hay empate entre uno anterior y uno posterior, gana el anterior). Cada fila del registro de tarifas abajo indica si es `propia` o `heredada de <Mes Año>`.
- **Ventana de ingesta:** Ene'25 y Jun'25 no se ingieren (fuera del rango de SP publicados, que arranca en Jul'25; Ene'25 además usa un layout distinto).
- **Relación con SP:** la métrica USD/SP se calcula en el dashboard cruzando este log con `log_performance_desarrollo.md` por año × mes × espacio. Si un espacio no tiene SP publicados ese mes (hoy: OB y SER), USD/SP da **0** — no se oculta ni se propaga a otros espacios.

## Registro de lotes ingeridos

{chr(10).join(lots_tbl)}

## Registro de tarifas por perfil

{chr(10).join(tarifas)}

## Resumen mensual (horas / USD de desarrollo)

{chr(10).join(resumen)}

## Datos — detalle año × mes × espacio

{chr(10).join(detalle)}
"""
    LOG_COST_MD.write_text(md, encoding="utf-8")


def nearest_period(target, known_periods):
    """Mes conocido (con tarifa propia) más cercano a `target`; empate -> gana el anterior."""
    known_periods = [p for p in known_periods if p != target]
    if not known_periods:
        return None
    def dist(p):
        return abs((p[0] * 12 + p[1]) - (target[0] * 12 + target[1]))
    return min(known_periods, key=lambda p: (dist(p), 0 if p <= target else 1))


def resolve_stock_costs(stock_segments, existing_rate_rows):
    """Toma los segmentos (rows, meta) de formato 3 leídos de raw/ + el registro de
    tarifas ya persistido, y devuelve, por período (año, mes_num):
      {"anio", "mes_num", "mes", "skip", "espacios": {esp: {"horas","usd"}},
       "origen_label", "missing_perfiles": set, "perfiles_usados": {perfil: usd_hora}}
    No escribe nada — lo usan tanto `inspect` (solo imprime) como `ingest` (además persiste)."""
    known_propia = {}
    for r in existing_rate_rows:
        if r["origen"] == "propia":
            known_propia.setdefault((r["anio"], r["mes_num"]), {})[r["perfil"]] = r["usd_hora"]

    by_period = defaultdict(list)   # (anio,mes_num) -> [row,...]
    meta_by_period = {}
    for rows, meta in stock_segments:
        if meta["kind"] != "cost":
            continue
        period = (meta["anio"], meta["mes_num"])
        by_period[period].extend(rows)
        meta_by_period.setdefault(period, []).append(meta)
        if meta["own_rates"]:
            known_propia[period] = dict(meta["own_rates"])  # el propio del batch pisa al persistido

    results = {}
    for period, rows in by_period.items():
        anio, mes_num = period
        skip = any(m["skip"] for m in meta_by_period[period])
        if period in known_propia and any(m["own_rates"] for m in meta_by_period[period]):
            effective, origen_label = known_propia[period], "propia"
        else:
            np_ = nearest_period(period, known_propia.keys())
            if np_:
                effective = known_propia[np_]
                origen_label = f"heredada de {MESES[np_[1]-1]} {np_[0]}"
            else:
                effective, origen_label = {}, "sin tarifa conocida (USD=0)"

        espacios, perfiles_usados, missing = defaultdict(lambda: [0.0, 0.0]), {}, set()
        for row in rows:
            r = effective.get(row["tipo"])
            if r is None:
                missing.add(row["tipo"])
                r = 0.0
            perfiles_usados[row["tipo"]] = r
            espacios[row["espacio"]][0] += row["horas"]
            espacios[row["espacio"]][1] += row["horas"] * r
        results[period] = {
            "anio": anio, "mes_num": mes_num, "mes": MESES[mes_num - 1], "skip": skip,
            "espacios": {e: {"horas": round(h, 2), "usd": round(u, 2)} for e, (h, u) in espacios.items()},
            "origen_label": origen_label, "missing_perfiles": missing, "perfiles_usados": perfiles_usados,
        }
    return results


# --- Dashboard --------------------------------------------------------------
def write_dashboard(records, cost_records, sla_records, today_ddmmyyyy):
    tpl = TEMPLATE.read_text(encoding="utf-8")
    subtitle = (f"Tickets, story points y costo de desarrollo <strong>publicados en producción</strong> · "
                f"Fuente: reporte mensual del PM (Jira) + stock de horas de Fintexa + backfills puntuales · "
                f"Última ingesta: {today_ddmmyyyy} ({coverage_str(records)})")
    out = tpl.replace("__DATA_JSON__", json.dumps(records, ensure_ascii=False, separators=(",", ":")))
    out = out.replace("__COSTOS_JSON__", json.dumps(cost_records, ensure_ascii=False, separators=(",", ":")))
    out = out.replace("__SLA_JSON__", json.dumps(sla_records, ensure_ascii=False, separators=(",", ":")))
    out = out.replace("__SUBTITLE__", subtitle)
    assert "__DATA_JSON__" not in out and "__COSTOS_JSON__" not in out and "__SLA_JSON__" not in out and "__SUBTITLE__" not in out
    DASHBOARD.write_text(out, encoding="utf-8")


# --- Comandos ----------------------------------------------------------------
def report_meta(meta):
    hoja = f" · hoja: {meta['sheet']}" if "sheet" in meta else ""
    print(f"  formato: {meta['format']}{hoja} · filas: {meta['rows']}")
    if meta["extras"]:
        print(f"  [NUEVAS COLUMNAS no contempladas — evaluar si aportan métricas nuevas]: {meta['extras']}")
    for w in meta["warnings"]:
        print(f"  [WARN] {w}")


def cmd_inspect():
    seed_staging_from_mirror()
    excels = find_excels()
    csvs = find_csvs()
    if not excels and not csvs:
        sys.exit("[ABORT] No hay .xlsx/.xlsm/.csv en raw/ (se ignoran locks ~$). Nada que ingestar.")

    if csvs:
        print("== CSV de SLA Highest ==")
        existing_sla, _ = parse_sla_log()
        today_dt = datetime.now()
        for p in csvs:
            fmt = sniff_csv_format(p)
            if fmt != "sla_highest":
                sys.exit(f"[ABORT] {p.name}: ningún formato CSV conocido matchea este header (hoy solo se reconoce 'SLA Highest', columnas Clave/Estado/Creada + tiempo-en-estado Asignado/Backlog/Finalizada). No se adivina el mapeo — revisar con el usuario.")
            print(f"-- {p.name} --")
            tickets, meta = read_sla_csv(p, today_dt)
            report_meta(meta)
            overlap = set(tickets) & set(existing_sla)
            nuevos = len(set(tickets) - set(existing_sla))
            print(f"  {nuevos} tickets nuevos respecto del log · {len(overlap)} ya existían")
        print()

    delivery_segments, stock_segments = [], []
    for fi, p in enumerate(excels):
        print(f"== {p.name} ==")
        for records, meta in sniff_and_read(p):
            report_meta(meta)
            if meta["kind"] == "delivery":
                delivery_segments.append((records, meta))
                ctrl = defaultdict(lambda: [0, 0.0])
                for r in records:
                    ctrl[(r["anio"], r["mes"], r["espacio"])][0] += r["tickets"]
                    ctrl[(r["anio"], r["mes"], r["espacio"])][1] += r["sp"]
                for (a, m, e) in sorted(ctrl, key=lambda k: (k[0], MESES.index(k[1]), k[2])):
                    tk, sp = ctrl[(a, m, e)]
                    print(f"  {m} {a} {e}: {tk} tk / {fnum(sp)} SP")
                existing, _ = parse_log()
                overlap = {(r["anio"], r["mes"], r["espacio"]) for r in existing} & set(ctrl)
                if overlap:
                    print(f"  [OVERLAP] combos ya en el log de delivery que este Excel PISARÍA: {sorted(overlap, key=lambda k: (k[0], MESES.index(k[1]), k[2]))}")
            else:
                stock_segments.append((records, meta, fi))

    if stock_segments:
        # Mismo criterio que cmd_ingest: si dos archivos traen el mismo período,
        # gana el último procesado (orden alfabético) — no se suman ambos.
        last_file_for_period = {}
        for _, meta, fi in stock_segments:
            last_file_for_period[(meta["anio"], meta["mes_num"])] = fi
        deduped = [(r, m) for r, m, fi in stock_segments
                   if fi == last_file_for_period[(m["anio"], m["mes_num"])]]
        _, _, existing_rates = parse_cost_log()
        resolved = resolve_stock_costs(deduped, existing_rates)
        existing_cost, _, _ = parse_cost_log()
        print("\n== Costos de desarrollo (formato stock de horas) — resuelto con carry-forward de tarifas ==")
        for period in sorted(resolved, key=lambda p: (p[0], p[1])):
            info = resolved[period]
            tag = " [FUERA DE VENTANA — no se ingiere]" if info["skip"] else ""
            print(f"  {info['mes']} {info['anio']}  (tarifa: {info['origen_label']}){tag}")
            for e in sorted(info["espacios"]):
                d = info["espacios"][e]
                print(f"    {e}: {fnum(d['horas'])} hs / ${fnum(d['usd'])}")
            if info["missing_perfiles"]:
                print(f"    [WARN] perfil(es) sin tarifa (USD=0 para esas horas): {sorted(info['missing_perfiles'])}")
        overlap = {(r["anio"], r["mes"], r["espacio"]) for r in existing_cost} & {
            (info["anio"], info["mes"], e) for info in resolved.values() if not info["skip"] for e in info["espacios"]
        }
        if overlap:
            print(f"  [OVERLAP] combos ya en el log de costos que esta ingesta PISARÍA: {sorted(overlap, key=lambda k: (k[0], MESES.index(k[1]), k[2]))}")


def cmd_ingest():
    seed_staging_from_mirror()
    excels = find_excels()
    csvs = find_csvs()
    if not excels and not csvs:
        sys.exit("[ABORT] No hay .xlsx/.xlsm/.csv en raw/ (se ignoran locks ~$). Nada que ingestar.")
    today = date.today().isoformat()
    today_dd = date.today().strftime("%d/%m/%Y")
    today_dt = datetime.now()

    existing, lots = parse_log()
    cost_existing, cost_lots, rate_rows = parse_cost_log()
    sla_existing, sla_lots = parse_sla_log()

    if csvs:
        for p in csvs:
            fmt = sniff_csv_format(p)
            if fmt != "sla_highest":
                sys.exit(f"[ABORT] {p.name}: ningún formato CSV conocido matchea este header. No se adivina el mapeo — revisar con el usuario.")
            print(f"== Ingesta SLA: {p.name} ==")
            tickets, meta = read_sla_csv(p, today_dt)
            report_meta(meta)
            sla_existing.update(tickets)
            dest = f"`4_archivos/historial_raw/{today[:7]}_sla_highest_jira/`"
            lot_row = [today, f"`{p.name}`", sla_coverage_str(list(tickets.values())), f"{len(tickets)} tickets", dest]
            sla_lots = [l for l in sla_lots if l[1] != lot_row[1]] + [lot_row]
        write_sla_log(sla_existing, sla_lots, today)

    # Pass 1: leer todos los archivos, mergear delivery en el momento (semántica
    # "último archivo procesado pisa" ya correcta al ser secuencial), y acumular los
    # segmentos de costo de TODOS los archivos para resolverlos juntos en el Pass 2
    # (una tarifa "propia" encontrada en un archivo posterior debe poder heredarse
    # hacia atrás a un archivo ya procesado — no puede resolverse archivo por archivo).
    all_stock_segs = []  # (records, meta, file_index, filename)
    for fi, p in enumerate(excels):
        print(f"== Ingesta: {p.name} ==")
        segments = sniff_and_read(p)
        delivery_segs = [(r, m) for r, m in segments if m["kind"] == "delivery"]
        stock_segs = [(r, m) for r, m in segments if m["kind"] == "cost"]

        for new_records, meta in delivery_segs:
            report_meta(meta)
            combos = {(r["anio"], r["mes"], r["espacio"]) for r in new_records}
            old_ctrl = defaultdict(lambda: [0, 0.0])
            for r in existing:
                if (r["anio"], r["mes"], r["espacio"]) in combos:
                    old_ctrl[(r["anio"], r["mes"], r["espacio"])][0] += r["tickets"]
                    old_ctrl[(r["anio"], r["mes"], r["espacio"])][1] += r["sp"]
            for c in sorted(combos & set(old_ctrl), key=lambda k: (k[0], MESES.index(k[1]), k[2])):
                new_tk = sum(r["tickets"] for r in new_records if (r["anio"], r["mes"], r["espacio"]) == c)
                new_sp = sum(r["sp"] for r in new_records if (r["anio"], r["mes"], r["espacio"]) == c)
                tag = "sin cambios" if (old_ctrl[c][0], round(old_ctrl[c][1], 2)) == (new_tk, round(new_sp, 2)) \
                    else f"CAMBIÓ: era {old_ctrl[c][0]} tk / {fnum(old_ctrl[c][1])} SP → ahora {new_tk} tk / {fnum(new_sp)} SP"
                print(f"  [REEMPLAZA] {c[1]} {c[0]} {c[2]} — {tag}")

            existing = [r for r in existing if (r["anio"], r["mes"], r["espacio"]) not in combos] + new_records
            tk = sum(r["tickets"] for r in new_records)
            sp = sum(r["sp"] for r in new_records)
            slug = "backfill_historico" if meta["format"] == "agregado-por-version" else "reporte_pm_metricas_publicadas"
            dest = f"`4_archivos/historial_raw/{today[:7]}_{slug}/`"
            lot_row = [today, f"`{p.name}`", coverage_str(new_records), str(tk), fnum(sp), dest]
            lots = [l for l in lots if l[1] != lot_row[1]] + [lot_row]

        for records, meta in stock_segs:
            report_meta(meta)
            all_stock_segs.append((records, meta, fi, p.name))

    # Pass 2: si dos archivos traen el mismo período, gana el último procesado (mismo
    # criterio que delivery y que cmd_inspect) — no se suman ambos.
    if all_stock_segs:
        last_file_for_period = {}
        for _, meta, fi, _ in all_stock_segs:
            last_file_for_period[(meta["anio"], meta["mes_num"])] = fi
        deduped = [(r, m, fname) for r, m, fi, fname in all_stock_segs
                   if fi == last_file_for_period[(m["anio"], m["mes_num"])]]
        resolved = resolve_stock_costs([(r, m) for r, m, _ in deduped], rate_rows)
        file_of_period = {(m["anio"], m["mes_num"]): fname for _, m, fname in deduped}

        new_cost_rows = []
        per_file_totals = defaultdict(lambda: [0.0, 0.0, []])  # fname -> [horas, usd, records]
        for period, info in sorted(resolved.items()):
            fname = file_of_period[period]
            if info["skip"]:
                print(f"  [{fname}] {info['mes']} {info['anio']}: fuera de ventana de SP — NO se ingiere.")
                continue
            print(f"  [{fname}] {info['mes']} {info['anio']}  (tarifa: {info['origen_label']})")
            for e, d in sorted(info["espacios"].items()):
                print(f"    [REEMPLAZA] {e}: {fnum(d['horas'])} hs / ${fnum(d['usd'])}")
                rec = {"anio": info["anio"], "mes": info["mes"], "espacio": e, "horas": d["horas"], "usd": d["usd"]}
                new_cost_rows.append(rec)
                per_file_totals[fname][0] += d["horas"]
                per_file_totals[fname][1] += d["usd"]
                per_file_totals[fname][2].append(rec)
            if info["missing_perfiles"]:
                print(f"    [WARN] perfil(es) sin tarifa (USD=0 para esas horas): {sorted(info['missing_perfiles'])}")
            rate_rows = [r for r in rate_rows if (r["anio"], r["mes_num"]) != period]
            rate_rows.extend({"anio": info["anio"], "mes": info["mes"], "mes_num": info["mes_num"],
                              "perfil": perfil, "usd_hora": usd_hora, "origen": info["origen_label"]}
                             for perfil, usd_hora in info["perfiles_usados"].items())

        if new_cost_rows:
            touched_combos = {(r["anio"], r["mes"], r["espacio"]) for r in new_cost_rows}
            cost_existing = [r for r in cost_existing if (r["anio"], r["mes"], r["espacio"]) not in touched_combos] + new_cost_rows
            dest = f"`4_archivos/historial_raw/{today[:7]}_backfill_stock_horas_fintexa/`"
            for fname, (tot_h, tot_u, recs) in per_file_totals.items():
                lot_row = [today, f"`{fname}`", coverage_str(recs), fnum(tot_h), fnum(tot_u), dest]
                cost_lots = [l for l in cost_lots if l[1] != lot_row[1]] + [lot_row]

    existing = sort_records(existing)
    write_log(existing, lots, today)
    cost_existing = sort_cost_records(cost_existing)
    write_cost_log(cost_existing, cost_lots, rate_rows, today)
    sla_sorted = sorted(sla_existing.values(), key=lambda t: (t["creada"], t["clave"]))
    write_dashboard(existing, cost_existing, sla_sorted, today_dd)

    print("\n== Estado acumulado post-ingesta (delivery) ==")
    for line in build_resumen(existing):
        print("  " + line)
    print("\n== Estado acumulado post-ingesta (costos) ==")
    for line in build_cost_resumen(cost_existing):
        print("  " + line)
    if sla_sorted:
        print("\n== Estado acumulado post-ingesta (SLA Highest) ==")
        for line in build_sla_resumen(sla_sorted):
            print("  " + line)
        print(f"OK -> {LOG_SLA_MD.relative_to(REPO)}")
    print(f"\nOK -> {LOG_MD.relative_to(REPO)}")
    print(f"OK -> {LOG_COST_MD.relative_to(REPO)}")
    print(f"OK -> {DASHBOARD.relative_to(REPO)}")
    print("PENDIENTE PARA CLAUDE: verificar dashboard, rotar raw/ -> historial, changelog, gaps si hubo warnings, git push.")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "inspect":
        cmd_inspect()
    elif cmd == "ingest":
        cmd_ingest()
    else:
        sys.exit("uso: python pipeline.py [inspect|ingest]")
